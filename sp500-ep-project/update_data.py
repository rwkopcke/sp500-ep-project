'''This program reads selected data from S&P, sp-500-eps-est.xlsx
        https://www.spglobal.com/spdji/en/search/?query=index+earnings&activeTab=all
   and from the 10-year TIPS rate from FRED: 
        https://fred.stlouisfed.org/series/DFII10
   It writes these data as polars dataframes to .parquet files
        and writes a record of the files that it has read and writen
        as a dictionary to a .json file
   The polars dataframes contain the latest projections of earnings for the
   S&P500 within each quarter since late 2017. A separate polars dataframe
   containsthe actual earnings and the value of the index for each quarter 
   beginning in 1988. This dataframe also contains actual values for 
   operating margins, revenues, book values, dividends, and other 
   actual data reported by S&P, plus actual values for the 10-year TIPS.
   
   The addresses of documents for this project appear in this program's 
   project directory: S&P500_PE/sp500_pe/__init__.py
'''

import sys
import gc

import polars as pl
import json
from openpyxl import load_workbook

import paths as sp
import func_module.helper_func as hp
import func_module.read_data_func as rd

#######################  Parameters  ##################################

# FOR DEBUGGING -- run program, with the value True
# True allows inspection of dfs & aborts writing new files
HALT_PROCESS = True

# data from "ESTIMATES&PEs" wksht
RR_COL_NAME = 'real_int_rate'
YR_QTR_NAME = 'yr_qtr'
PREFIX_OUTPUT_FILE_NAME = 'sp-500-eps-est'
EXT_OUTPUT_FILE_NAME = '.parquet'

SHT_EST_NAME = "ESTIMATES&PEs"
COLUMN_NAMES = ['date', 'price', 'op_eps', 'rep_eps',
                'op_p/e', 'rep_p/e', '12m_op_eps', '12m_rep_eps']
PROJ_COLUMN_NAMES = ['date', 'op_eps', 'rep_eps',
                     'op_p/e', 'rep_p/e', '12m_op_eps', '12m_rep_eps']

SHT_QTR_NAME = "QUARTERLY DATA"
COLUMN_NAMES_QTR = ['date', 'div_ps', 'sales_ps',
                    'bk_val_ps', 'capex_ps', 'divisor']

SHT_IND_NAME = 'SECTOR EPS'

# all search (row or col) "keys" should be None or lists

SHT_EST_DATE_PARAMS = {
    'date_keys' : ['Date', 'Data as of the close of:'],
    'value_col_1' : 'D',
    'date_key_2' : ['ACTUALS'],
    'value_col_2' : 'B',
    'column_names' : COLUMN_NAMES
}

SHT_HIST_PARAMS = {
    'act_key' : ['ACTUALS', 'Actuals'],
    'first_col' : 'A',
    'last_col' : 'J',
    'skip_col' : [4, 7],
    'column_names' : COLUMN_NAMES
}

MARG_KEY = 'QTR'
SHT_BC_MARG_PARAMS = {
    'row_key': [MARG_KEY],
    'first_col': 'A',
    'stop_col_key': None,
    'stop_row_data_offset': 4,
    'yr_qtr_name': YR_QTR_NAME
}

NUM_BASE_IND_NAMES = 48  # 12 "industries", 4 size indexes
IND_KEY = 'INDEX NAME'
SHT_BC_IND_PARAMS = {
    'row_key': [IND_KEY],
    'first_col': 'A',
    'stop_col_key': None,
    'start_row_data_offset': 2,
    'stop_row_data_offset': 107,
    'num_inds': NUM_BASE_IND_NAMES,
    'yr_qtr_name': YR_QTR_NAME
}

SHT_QTR_PARAMS = {
    'act_key' : ['END'],
    'first_col' : 'A',
    'last_col' : 'I',
    'skip_col' : [2, 3, 7],
    'column_names' : COLUMN_NAMES_QTR
}

SHT_EST_PROJ_DATE_PARAMS = {
    'date_keys' : ['Date', 'Data as of the close of:'],
    'value_col_1' : 'D', 
    'date_key_2' : None, 
    'value_col_2' : None,
    'column_names' : None,
    'include_prices' : False
}

SHT_EST_PROJ_PARAMS = {
    'act_key' : ['ESTIMATES'],
    'first_col' : 'A',
    'last_col' : 'J',
    'skip_col' : [1, 4, 7],
    'column_names' : PROJ_COLUMN_NAMES
}

SHT_FRED_PARAMS = {
    'first_row': 12,
    'col_1': 'A',
    'col_2': 'B',
    'yr_qtr_name': YR_QTR_NAME,
    'rr_col_name': RR_COL_NAME
}


#######################  MAIN Function  ###############################

def update_data_files():
    '''create or update earnings, p/e, and margin data
       from 'sp-500-eps-est ...' files
    '''
    

# ++++++  PRELIMINARIES +++++++++++++++++++++++++++++++++++++++++++++++
# load file containing record_dict: record of files seen previously
#   if record_dict does not exist, create an empty dict to initialize
    if sp.RECORD_DICT_ADDR.exists():
        with sp.RECORD_DICT_ADDR.open('r') as f:
            record_dict = json.load(f)
        print('\n============================================')
        print(f'Read record_dict from: \n{sp.RECORD_DICT_ADDR}')
        print('============================================\n')
        
        # backup record_dict
        with sp.BACKUP_RECORD_DICT_ADDR.open('w') as f:
            json.dump(record_dict, f)
        print('============================================')
        print(f'Wrote record_dict to: \n{sp.BACKUP_RECORD_DICT_ADDR}')
        print('============================================\n')
        
    else:
        print('\n============================================')
        print(f'No record dict file found at: \n{sp.RECORD_DICT_ADDR}')
        print(f'Initialized record_dict with no entries')
        print('============================================\n')
        record_dict = {'sources': {'s&p': '',
                                   'tips': ''},
                       'latest_used_file': "",
                       'proj_yr_qtrs' : [],
                       'prev_used_files': [],
                       'output_proj_files': [],
                       'prev_files': []}
    
    # ensure that recorded sources are current
    record_dict['sources']['s&p'] = sp.SP_SOURCE
    record_dict['sources']['tips'] = sp.REAL_RATE_SOURCE
        
# create list of earnings input files not previously seen
# and add them to 'prev_files'
    prev_files_set = set(record_dict['prev_files'])
    
    new_files_set = \
        set(str(f.name) 
            for f in sp.INPUT_DIR.glob('sp-500-eps*.xlsx'))
    
    new_files_set = new_files_set - prev_files_set
    
    # if no new data, print alert and exit
    if len(new_files_set) == 0:
        print('\n============================================')
        print(f'No new files in {sp.INPUT_DIR}')
        print('All files have been read previously')
        print('============================================\n')
        sys.exit()
        
# there is new data, add new files to historical record
    record_dict['prev_files'] \
        .extend(list(new_files_set)) \
        .sort(reverse= True)

# find the latest new file for each quarter (agg(sort).last)
    data_df = pl.DataFrame(list(new_files_set), 
                          schema= ["new_files"],
                          orient= 'row')\
                .with_columns(pl.col('new_files')
                            .map_batches(hp.string_to_date)
                            .alias('date'))\
                .with_columns(pl.col('date')
                            .map_batches(hp.date_to_year_qtr)
                            .alias('yr_qtr'))\
                .group_by('yr_qtr')\
                .agg([pl.all().sort_by('date').last()])\
                .sort(by= 'yr_qtr') 
    
    files_to_archive = list(new_files_set)
    del new_files_set
    gc.collect()

# combine with prev_files where new_files has larger date for year_qtr
# (new files can update and replace prev files for same year_qtr)
# new_files has only one file per quarter -- no need for group_by
    prev_used = record_dict['prev_used_files']
    if len(prev_used) > 0:
        used_df = pl.DataFrame(prev_used, 
                               schema= ['used_files'],
                               orient= 'row')\
                .with_columns(pl.col('used_files')
                            .map_batches(hp.string_to_date)
                            .alias('date'))\
                .with_columns(pl.col('date')
                            .map_batches(hp.date_to_year_qtr)
                            .alias('yr_qtr'))
                
    # update used_files, a join with new files
        # 1st filter removes yr_qtr rows that have no dates in data_df
        # 2nd filter keeps only the rows with new data
    # after renaming, ensures that 'date' ref only files with new data
    # and proj_to_delete ref only files that are null or are superceded
        used_df = used_df.join(data_df,
                               on= 'yr_qtr',
                               how= 'full',
                               coalesce= True)\
                         .filter(pl.col('date_right').is_not_null())\
                         .filter(((pl.col('date').is_null()) | 
                                  (pl.col('date') <
                                   pl.col('date_right'))))\
                         .rename({'used_files' : 'proj_to_delete'})\
                         .drop(['date'])\
                         .rename({'date_right': 'date'})\
                         .sort(by= 'yr_qtr')
                         
        # remove old files from record_dict lists
        #   prev_used_files (.xlsx) & output_proj_files (.parquet)
        # remove the output .parquet file from output_proj_dir
        files_to_remove_list = \
            pl.Series(used_df.select(pl.col('proj_to_delete'))\
                             .filter(pl.col('proj_to_delete')
                                    .is_not_null()))\
                            .to_list()
                            
        for file in files_to_remove_list:
            record_dict['prev_used_files'].remove(file)
            file_list = file.split(" ", 1)
            proj_file = \
                f'{file_list[0]} {file_list[1]
                                    .replace(' ', '-')
                                    .replace('.xlsx', '.parquet')}'
            record_dict['output_proj_files'].remove(proj_file)
            # using Path() object
            address_proj_file = sp.OUTPUT_PROJ_DIR / proj_file
            if address_proj_file.exists():
                address_proj_file.unlink()
                print('\n============================================')
                print(f'Removed {proj_file} from: \n{sp.OUTPUT_PROJ_DIR}')
                print(f'Found file with more recent date for the quarter')
                print('============================================\n')
            else:
                print('\n============================================')
                print(f"WARNING")
                print(f"Tried to remove: \n{address_proj_file}")
                print(f'Address does not exist')
                print('============================================\n') 
                
    # when len(prev_used) == 0
    else:
        used_df = data_df

    del data_df
    gc.collect()           
    
    # add dates of projections and year_qtr to record_dict
    # https://www.rhosignal.com/posts/polars-nested-dtypes/   pl.list explanation
    # https://www.codemag.com/Article/2212051/Using-the-Polars-DataFrame-Library
    # pl.show_versions()

# files with new data: files_to_read_list, which is
# also used below in update projection files section
    files_to_read_list = \
        pl.Series(used_df.select('new_files')).to_list()
            
    # add dates of projections and year_qtr to record_dict
    record_dict['prev_used_files'] \
        .extend(files_to_read_list) \
        .sort(reverse= True)
        
    record_dict['proj_yr_qtrs']= \
        hp.date_to_year_qtr(
                hp.string_to_date(record_dict['prev_used_files'])
            ).to_list()
    # most recent is first
    record_dict["latest_used_file"] = record_dict['prev_used_files'][0]

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++              
## +++++  fetch the historical data  +++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++ 

    print('\n================================================')
    print(f'Updating historical data from: {record_dict["latest_used_file"]}')
    print(f'in directory: \n{sp.INPUT_DIR}')
    print('================================================\n')
    
## REAL INTEREST RATES, eoq, from FRED DFII10
    active_workbook = load_workbook(filename= sp.INPUT_RR_ADDR,
                                    read_only= True,
                                    data_only= True)
    active_sheet = active_workbook.active
    real_rt_df = rd.fred_reader(active_sheet,
                                **SHT_FRED_PARAMS)
    
## HISTORICAL DATA from existing .parquet file
    latest_file_addr = sp.INPUT_DIR / record_dict["latest_used_file"]
    
    # load s&p workbook, contains the most recent update to hist data
    active_workbook = load_workbook(filename= latest_file_addr,
                                    read_only= True,
                                    data_only= True)
    
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # actual df from parquet file
    # fetch date and new data from new excel file
    # concat new with old
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    # most recent date and prices
    active_sheet = active_workbook[SHT_EST_NAME]
    name_date, actual_df = rd.read_sp_date(active_sheet, 
                                           **SHT_EST_DATE_PARAMS,
                                           include_prices= True)
    
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# actual df should have only prices for recent quarters w/o historical data
# if len actual_df is zero, stop
# otherwise, update historical_df prices for those quarters?
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    # load historical data, if updates are available
    df = rd.sp_loader(active_sheet,
                      **SHT_HIST_PARAMS)
    
    # if any date is None, halt
    if (name_date is None or
        any([item is None
            for item in actual_df['date']])):
        
        print('\n============================================')
        print(f'Abort using {latest_file_addr} \nmissing history date')
        print(f'Name_date: {name_date}')
        print(actual_df['date'])
        print('============================================\n')
        sys.exit()
        
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# actual_df (YR_QTR_NAME) - existing historical_df(YR_QTR_NAME) parquet
# -> new quarters to be added to existing historical_df
# if nil, skip -> read the new projections
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        
    actual_df = pl.concat([actual_df, df], how= "diagonal")\
                  .with_columns(pl.col('date')
                        .map_batches(hp.date_to_year_qtr)
                        .alias(YR_QTR_NAME))
                  
    # merge real_rates with p and e history
    actual_df = actual_df.join( 
            real_rt_df, 
            how="left", 
            on=[YR_QTR_NAME],
            coalesce= True)
    
    del real_rt_df
    del df
    gc.collect()
        
## MARGINS
    margins_df = rd.margin_loader(active_sheet,
                                  **SHT_BC_MARG_PARAMS)
    
    # merge margins with previous data
    actual_df = actual_df.join(margins_df, 
                               how="left", 
                               on= YR_QTR_NAME,
                               coalesce= True)
    
    del margins_df
    gc.collect()

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    '''
## INDUSTRIAL DATA
    ind_df = rd.industry_loader(active_sheet,
                                **SHT_BC_IND_PARAMS)
    
    actual_df = actual_df.join(
        ind_df,
        how= 'left',
        on= YR_QTR_NAME,
        coalesce= True
    )
    
    del ind_df
    gc.collect()
    '''
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        
## QUARTERLY DATA
    active_sheet = active_workbook[SHT_QTR_NAME]

    qtrly_df = rd.sp_loader(active_sheet, 
                              **SHT_QTR_PARAMS)\
                 .with_columns(pl.col('date')
                            .map_batches(hp.date_to_year_qtr)
                            .alias(YR_QTR_NAME))
    
    # merge qtrly with previous data
    actual_df = actual_df.join(qtrly_df,  
                               how= "left", 
                               on= [YR_QTR_NAME],
                               coalesce= True)
    
    del qtrly_df
    gc.collect()

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++ update projection files +++++++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    # ordinarily a very short list
    # loop through files_to_read, fetch projections of earnings for each date
    failure_to_read_lst = []
    for file in files_to_read_list:
        # echo file name and address to console
        active_workbook = load_workbook(filename= sp.INPUT_DIR / file,
                                        read_only= True,
                                        data_only= True)
        active_sheet = active_workbook[SHT_EST_NAME]
        print(f'\n input file: {file}')    
        
# projections of earnings
        # read date of projection, no prices or other data
        name_date, _ = \
            rd.read_sp_date(active_sheet, 
                            **SHT_EST_PROJ_DATE_PARAMS)
        name_date = name_date.date()
    
        # load projections for the date
        proj_df = rd.sp_loader(active_sheet, 
                               **SHT_EST_PROJ_PARAMS)\
                    .with_columns(pl.col('date')
                        .map_batches(hp.date_to_year_qtr)
                        .alias(YR_QTR_NAME))

        # if any date is None, abort and continue
        if (name_date is None or
            any([item is None
                for item in proj_df['date']])):
            print('\n============================================')
            print('In main(), projections:')
            print(f'Skipped sp-500 {name_date} missing projection date')
            print('============================================\n')
            failure_to_read_lst.append(file)
            continue
        
############
        if HALT_PROCESS:
            print('\n============================================')
            print('HALT_PROCESS is True:')
            print('No data files have been written. End process.')
            print('============================================\n')
            sys.exit()
############

## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++ write files +++++++++++++++++++++++++++++++++++++++++++++++++++++++
## +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

## +++++  write proj_df  ++++++++++++++++++++++++++++++++++++++++++++++++++
        output_file_name = \
            f'{PREFIX_OUTPUT_FILE_NAME} {name_date}{EXT_OUTPUT_FILE_NAME}'
        record_dict['output_proj_files'].append(output_file_name)
        output_file_address = sp.OUTPUT_PROJ_DIR / output_file_name
        print(f'output file: {output_file_name}')
        
        with output_file_address.open('w') as f:
            proj_df.write_parquet(f)
            
## +++++ write history file ++++++++++++++++++++++++++++++++++++++++++++
    # move any existing hist file in output_dir to backup
    if sp.OUTPUT_HIST_ADDR.exists():
        sp.OUTPUT_HIST_ADDR.rename(sp.BACKUP_HIST_ADDR)
        print('\n============================================')
        print(f'Moved history file from: \n{sp.OUTPUT_HIST_ADDR}')
        print(f'to: \n{sp.BACKUP_HIST_ADDR}')
        print('============================================\n')
    else:
        print('\n============================================')
        print(f'Found no history file at: \n{sp.OUTPUT_HIST_ADDR}')
        print(f'Wrote no history file to: \n{sp.BACKUP_HIST_ADDR}')
        print('============================================\n')
        
    # write actual_df, the historical data, into the output file
    # save sp500_pe_dict to file
    with sp.OUTPUT_HIST_ADDR.open('w') as f:
        actual_df.write_parquet(f)
    print('\n============================================')
    print(f'Wrote history file to: \n{sp.OUTPUT_HIST_ADDR}')
    print('============================================\n')
            
## +++++ update archive ++++++++++++++++++++++++++++++++++++++++
    # archive all input files -- uses Path() variables
    # https://sysadminsage.com/python-move-file-to-another-directory/
    print('\n============================================')
    for file in files_to_archive:
        input_address = sp.INPUT_DIR / file
        if input_address.exists():
            input_address.rename(sp.ARCHIVE_DIR / file)
            print(f"Archived: {input_address}")
            
        else:
            print(f"\nWARNING")
            print(f"Tried: {input_address}")
            print(f'Address does not exist\n')
    print('============================================\n')
        
    sp.INPUT_RR_ADDR.rename(sp.ARCHIVE_DIR / sp.INPUT_RR_FILE)
    print('\n============================================')
    print(f"Archived: \n{sp.INPUT_RR_FILE}")
    print('============================================\n')
            
    # list should begin with most recent items
    # more efficient search for items to edit above
    record_dict['prev_files'].sort(reverse= True)
    record_dict['prev_used_files'].sort(reverse= True)
    record_dict['output_proj_files'].sort(reverse= True)
            
## store record_dict
    with sp.RECORD_DICT_ADDR.open('w') as f:
        json.dump(record_dict, f)
    print('\n====================================================')
    print('Saved record_dict to file')
    print(f'{sp.RECORD_DICT_ADDR}')
    print(f'\nlatest_used_file: {record_dict['latest_used_file']}\n')
    print(f'output_proj_files: \n{record_dict['output_proj_files'][:6]}\n')
    print(f'prev_used_files: \n{record_dict['prev_used_files'][:6]}\n')
    print(f'prev_files: \n{record_dict['prev_files'][:6]}\n')
    print(f'proj_yr_qtrs: \n{record_dict['proj_yr_qtrs'][:6]}\n')
    print('====================================================\n')
 
    print('\n====================================================')
    print('Retrieval is complete\n')
    
    n = len(files_to_read_list)
    m = len(failure_to_read_lst)
    print(f'{n} new input files read and saved')
    print(f'from {sp.INPUT_DIR}')
    print(f'  to {sp.OUTPUT_DIR}\n')
    print(f'{m} files not read and saved:\n')
    print(failure_to_read_lst)
    print('====================================================')

if __name__ == '__main__':
    update_data_files()