'''
   these are functions used by the update_data and display_data
   scripts and by the read_data_func functions
        
   access these values in other modules by
        import sp500_pe.helper_func as hp
'''
import sys

from datetime import datetime

import openpyxl.utils.cell
import polars as pl


def my_df_print(df):
    '''
        custom print df function to format data
    '''                                
    with pl.Config(
        tbl_cell_numeric_alignment="RIGHT",
        thousands_separator=",",
        float_precision=1
    ):
        print(df)
        
        
def dt_str_to_date(item):
    '''
        input either str or datetime obj
        if str, return datetime
    '''
    
    if isinstance(item, str):
        # fetch just the date component of str
        dt = item.split(" ")[0]
        dt = datetime.strptime(dt,'%m/%d/%Y')
    else:
        dt = item
    return dt
        

def string_to_date(series):
    '''
        receives pl.Series (col from df), str file names
        extracts the date string, "yyyy mm dd",
        returns the string to date object, as pl.Series
    '''
    # splits f_name into two two words,
    # the first and the rest
    # then splits the .xlsx from the rest
    return pl.Series([datetime.strptime((f_name.split(' ', 1)[1])
                                               .split('.', 1)[0], 
                                        '%Y %m %d').date()
                      for f_name in series])
    
    
def date_to_year_qtr(series):
    '''
        Receives pl.Series (col from df)
        Returns year_qtr string, yyyy-Qq, as pl.Series
    '''
    return pl.Series([f"{date.year}-Q{date_to_qtr(date)}"
            for date in series])
    
    
def date_to_qtr(date):
    '''
        Returns qtr number as string
    '''
    return f"{((date.month) - 1) // 3 + 1 }"


def is_quarter_4(series):
    '''
        returns bool: T if qtr == 4; else F
    '''
    return pl.Series([yq[-1] == '4'
                      for yq in series])
    
    
def yrqtr_to_yr(series):
    '''series of strings y-q in,
       series of strings y out
    '''
    return pl.Series([yq[:4]
                      for yq in series])
    

def find_key_row(wksht, search_col, start_row, key_values= None):
    '''
        for key_values (either None or a list),
        find cell containing (one of) the specified key(s)
        crawl down col A; return the row number of the first match
    '''
    
    # cap the number of rows to read
    max_to_read = wksht.max_row
    
    row_number = start_row
    while row_number < max_to_read:
        item = wksht[f'{search_col}{row_number}'].value
        
        if item_matches_key(item, key_values):
            return row_number
        row_number += 1                                                     
    return 0


def item_matches_key(item, keys):
    '''
        keys are either None or list of str
        check if item matches (one of) the key(s)
        return bool: T if match, F otherwise
    '''
    
    if (keys is None):
        # item is None? return T(match) or F(no match)
        return item is None
    
    # if item is not a str, it cannot match a key
    if not (type(item) is str):
        return False
    
    # at this point, keys must be a list of strings
    # if singleton, convert to list
    if isinstance(keys, str):
        keys = list(keys)
    
    # is item in keys?
    try:
        if all(isinstance(y, str) for y in keys):
            # T(match in keys) or F(no match in keys)
            return item in keys
    # in case iter (keys) creates an error
    except TypeError:
        pass
    print('\n============================================')
    print(f'In helper_func.py, item_matches_key(item, keys)')
    print(f'item: {item}')
    print(f'keys: {keys}, is not a list of strings')
    print('============================================\n')
    sys.exit()


def find_key_col(wksht, search_row, start_col, key_value= None):
    '''
        crawl along search_row to
        find cell containing the specified key,
        return col number of the first cell that matches
    '''
    
    # cap the number of rows to read
    max_to_read = wksht.max_column
    
    col_numb = start_col
    while col_numb < max_to_read:
        # convert col number to letter
        col_ltr = openpyxl.utils.cell.get_column_letter(col_numb)
        item = wksht[f'{col_ltr}{search_row}'].value
        
        if item_matches_key(item, key_value):
            break
        col_numb += 1
        
    if col_numb == max_to_read:
        return 0    
                                                                  
    return col_numb
    