'''
   these are functions used by the update_data and display_data
   scripts
        
   access these values in other modules by
        import sp500_pe.read_data_func as rd
'''

import sys

from openpyxl import load_workbook
import openpyxl.utils.cell
import polars as pl
import polars.selectors as cs

import func_module.helper_func as hp

def read_sp_date(wksht,
                 date_keys, value_col_1, 
                 date_key_2, value_col_2,
                 column_names, include_prices= False):
    '''
        fetch date of excel workbook from s&p
        fetch dates and prices that have occurred after
        the last reported set of financial data if
        include_prices= True
        return date in name_date
        (optional) return df with recent dates and prices
    '''
    
    if date_keys is None:
        print('\n============================================')
        print(f'Date keys are {date_keys} for {wksht}')
        print('============================================\n')
        sys.exit()
    
    # fetch row for latest date and price
    key_row = hp.find_key_row(wksht, 'A', 1, date_keys)

    if (key_row == 0):
        print('\n============================================')
        print(f'Found no {date_keys} in {wksht}')
        print('============================================\n')
        sys.exit()
        
    name_date = hp.dt_str_to_date(
                    wksht[f'{value_col_1}{key_row}'].value)
    
    # return without prices if include_prices is False
    if not include_prices:
        return [name_date, None]
        
    date_lst = []
    price_lst = []
    
    date_lst.append(name_date)
    name_date = name_date.date()  # value to return should be date()
    price_lst.append(wksht[f'{value_col_1}{key_row + 1}'].value)
    
    # fetch next date and price
    key_row = hp.find_key_row(wksht, 'A', key_row, date_key_2)
    
    if (key_row == 0):
        print('\n============================================')
        print(f'Found no {date_key_2} in {wksht}')
        print('============================================\n')
        sys.exit()
    
    date_lst.append(hp.dt_str_to_date(
        wksht[f'A{key_row - 2}'].value))
    price_lst.append(wksht[f'{value_col_2}{key_row -2}'].value)
    
    df = pl.DataFrame({
                column_names[0]: date_lst,
                column_names[1]: price_lst},
                schema= {column_names[0]: pl.Date, 
                            column_names[1]: pl.Float32})
    return [name_date, df]


def data_block_reader(wksht, start_row, stop_row,
                      first_col, last_col,
                      skip_col):
    """
    This function returns the block of data in a worksheet
    as a list of lists
    """
    
    # read a list of lists (rows)
    rng = wksht[f'{first_col}{start_row}:{last_col}{stop_row}']
    data = [[col_cell.value 
                for ind, col_cell in enumerate(row)
                if ind not in skip_col]
            for row in rng]
    
    return data
    

def sp_loader(wksht,
              num_rows_to_read,
              act_key, first_col, last_col,
              skip_col, column_names):
    '''
        read data from s&p excel workbook sheet
        that contains history for prices and earnings
        return df
    '''
    
    # fetch historical earnings data from wksht
    # fix the block of rows and cols that contain the data
    key_row = hp.find_key_row(wksht, 'A', 1, act_key)
    
    # first data row to read. Follows key_wrd row.
    start_row = 1 + key_row
    # last data row to read. 
    stop_row = start_row - 1 + num_rows_to_read
    
    # fetch the data from the block
    data = data_block_reader(wksht, start_row, stop_row,
                             first_col, last_col, skip_col)
    # iterate over rows to convert all str dates to datetime.date
    # row[0]: datetime or str, '%m/%d/%Y' is first 'word' in str
    for row in data:
        row[0] = hp.dt_str_to_date(row[0])
        
    df = pl.DataFrame(data, schema=column_names, orient="row")\
                .cast({cs.float(): pl.Float32,
                       cs.datetime(): pl.Date})
    return df


def margin_loader(wksht, row_key, first_col, stop_col_key,
                  stop_row_data_offset,
                  yr_qtr_name):
    
    '''
        read data from s&p excel worksheet
        that contains history for margins
        return df
    '''
    
    # find the rows with dates and data
    # start row contains dates
    start_row = hp.find_key_row(wksht, 'A', 1, row_key)
    
    stop_row_data = start_row + stop_row_data_offset
    
    # find last col with data
    last_col = -1 + hp.find_key_col(wksht, start_row, 
                                    2, stop_col_key)
    
    # some date info in start_row, some in data rows
    # block to collect: first_blk_col, start_row_data  
    #               to  last_blk_col, stop_row_data
    
     # list of lists for each row, including headers for cols (start_row)
    stop_col =  openpyxl.utils.cell.get_column_letter(last_col)
        
    data =  data_block_reader(wksht, start_row, stop_row_data,
                              first_col, stop_col, [])
     
    # data_values omits the first row (col headers) from data
    data_values = [row for row in data[1:]]
    
    # omit the * for 2008, take first entry (yr) in list data[0]
    col_names = [str(item).split('*')[0] for item in data[0]]
   
    # build "tall" 2-col DF with 'year_qtr' and 'margin'
    df = pl.DataFrame(data_values, schema= col_names,
                      orient= 'row')\
                .with_columns(pl.col('QTR')
                              .map_elements(lambda x: x.split(' ')[0],
                                            return_dtype= str))\
                .cast({cs.float(): pl.Float32})\
                .unpivot(index= 'QTR', variable_name='year')
            # index: names of cols to remain cols
            # variable_name: name of col to contain names of cols pivoted
    
    df = df.with_columns(
                pl.struct(['QTR', 'year'])\
                    .map_elements(lambda x: 
                                  f"{x['year']}-{x['QTR']}",
                                  return_dtype= pl.String)\
                    .alias(yr_qtr_name))\
            .drop(['year', 'QTR'])\
            .rename({'value': 'op_margin'})
    return df


def industry_loader(wksht, 
                    row_key, 
                    first_col, stop_col_key,
                    start_row_data_offset, stop_row_data_offset,
                    num_inds, yr_qtr_name):
    '''
        read data from s&p excel worksheet
        that contains history for industry data
        return df
    '''
    
    # find the rows with dates and data
    # start row contains dates
    start_row = hp.find_key_row(wksht, 'A', 1, row_key)
    
    start_row_data = start_row + start_row_data_offset
    stop_row_data = start_row + stop_row_data_offset
    
    # find last col with data
    last_col = -1 + hp.find_key_col(wksht, start_row, 
                                        2, stop_col_key)
    
    stop_col =  openpyxl.utils.cell\
        .get_column_letter(last_col)
    
    # block to collect: first_blk_col, start_row_data  
    #               to  last_blk_col, stop_row_data
    
    # fetch dates for the block from start_row only
    dates_raw = data_block_reader(wksht, start_row, start_row,
                              first_col, stop_col, [])
    
    # fetch 1st row to build the dates for the data
    # first 4 char in str are year; last char is qtr #
    dates = (f'{item[:4]}-Q{item[-1:]}'
             for item in dates_raw)
    
    # fetch the data
    # list of lists for each row
    data = data_block_reader(wksht, start_row_data, stop_row_data,
                              first_col, stop_col, [])
    # remove rows without data
    data = [row 
            for row in data
            if row[0] not in 
              ['', 'As Reported Earnings Per Share by Economic Sector']]
    
    # data values appear in 3rd through last cols of data
    data_values = [row[2:] for row in data]
    # current prices in 2nd col
    current_price_industries = [row[1] for row in data]
    # extract ind name from 1st col, without parenthetical details
    base_col_names= [row[0].split(' (')[0]
                     for row in data[:num_inds]]
    
    col_names = [
        *['op_E ' + item for item in base_col_names],
        *['rep_E ' + item for item in base_col_names]
    ]
   
    # build "tall" many-col DF with year_qtr and industry cols
    # cheap way to accomplish 'transpose' (orient= col)
    # each row in data_values is a col in df
    df = pl.DataFrame(data_values, 
                      schema= col_names,
                      orient= 'col')\
            .cast({cs.float(): pl.Float32})\
            .with_columns(pl.Series(dates)
                      .alias(yr_qtr_name))
            
    return df


def fred_reader(wksht, first_row, col_1, col_2,
                yr_qtr_name, rr_col_name):
    '''
        read data from FRED excel worksheet
        that contains history for real interest rates
        return df
    '''

    last_row = wksht.max_row
    data = data_block_reader(wksht, first_row, last_row,
                             col_1, col_2, [])
    
    df = pl.DataFrame(data, schema=['date', rr_col_name],
                      orient='row')\
           .with_columns(pl.col('date')
                        .map_batches(hp.date_to_year_qtr)
                        .alias(yr_qtr_name))\
           .group_by(yr_qtr_name)\
           .agg([pl.all().sort_by('date').last()])\
           .sort(by= yr_qtr_name)\
           .drop('date')\
           .cast({cs.datetime(): pl.Date,
                  cs.float(): pl.Float32})
    return df